#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Author   : FLS
# @Software : PyCharm
# @Time     : 2022/12/6 18:21
# @Project  : pycharmlearningproject
# @File     : openpyxl的跨文件复制sheet.py



'''
实现跨excel文件复制sheet的功能,默认条件是openpyxl是没有禁止跨文件复制的，只是会有一个断言.
'''

import openpyxl

def copy_excel_file(file1,file2,sheet1,sheet2):
    wb1=openpyxl.load_workbook(file1)
    wb2=openpyxl.load_workbook(file2)
    ws1=wb1[sheet1]
    ws2=wb2[sheet2]

    for i,row in enumerate(ws1.iter_rows()):
        for j,cell in enumerate(row):
            ws2.cell(row=i+1,column=j+1,value=cell.value)

    wb2.save(file2)

copy_excel_file('1.xlsx','2.xlsx','A','B')
