# /usr/bin/env python3
# -*- coding: utf-8 -*-
#
# Copyright (C) 2017-2020  The Project X-Ray Authors.
#
# Use of this source code is governed by a ISC-style
# license that can be found in the LICENSE file or at

import openpyxl



'''
层级：
workbook  工作簿，就是一个xlsl文件
worksheet  工作表  文件里面的一个表
rows  一行
column  一列
cell  一个单元格
'''


# 一、sheet级别的操作
# 指定sheet的操作，就两种
# wb=openpyxl.load_workbook('a.xlsx')
# ws1=wb['ccc']  # 名字
# ws2=wb.worksheets[0] # 顺序
# ws3=wb.get_sheet_by_name('ccc') # 过时
# wb.save('a.xlsx')

# sheet的增删改查
# wb=openpyxl.Workbook('a.xlsx')
# ws1=wb.create_sheet('c',0)
# ws1.title = 'ccc'
# ws1.sheet_properties.tabColor='1072BA'  # 设置sheet名的参数
# print(wb.sheetnames)
# for sheet in wb:
#     print(sheet.title)
# wb.save('a.xlsx')

# 复制sheet,原生只能在同一个wb内进行复制，不能复制图片和图表
# wb=openpyxl.load_workbook('a.xlsx',read_only=False)
# print(wb.sheetnames)
# ws1=wb['ccc']
# ws2=wb.copy_worksheet(from_worksheet=ws1)
# ws2.title="ddd"
# ws3=wb.worksheets[1]
# print(ws3.title)
# wb.remove(wb.worksheets[1])

# 删除sheet操作
# # del wb.worksheets[2]
# print(ws2.cell(1, 2).value)
# wb.save('a.xlsx')


# 保存成宏文件：
# wb = load_workbook('document.xltm', keep_vba=True)
# wb.save('new_document.xlsm')


# 保存成网络流文件
# from tempfile  import NamedTemporaryFile
# from openpyxl import Workbook
# wb=Workbook()
# with NamedTemporaryFile() as temp:
#     wb.save(temp.name)
#     temp.seek(0)
#     stream=temp.read()


#  打开操作
wb = openpyxl.load_workbook('shuju.xlsx')
print(wb.sheetnames)
print(wb.encoding)
ws=wb['BC']
print(ws.title)
print(ws)

#  工作簿。工作表、单元格的属性
# print(wb.sheetnames)
# print(wb.worksheets) # 所有的表名
#
# print(ws.title) # 当前表名
# print(wb.read_only) # 判断是否只读
# print(wb.encoding)
# print(wb.properties)  # 工作簿元数据

# print(ws.max_row)
# print(ws.max_column)


# 工作簿、表的操作。
# 复制工作表
# target=wb.copy_worksheet(ws)
# print(target)
# target.title='FLS'
# wb.save('shuju.xlsx')


# 读取单元格数据
# cells=ws['A:B]
# cells=ws['5:6:7:8:9:10:11:12:']

# # 获取区域单元格的的序列：
# for row in ws.iter_rows(min_row=2,max_row=10,min_col=2,max_col=3):
#     print(row)
#     for cell in row:
#         print(cell.value)

# ws.iter_cols(min_row=1,max_col=3,max_row=3)是同样的使用方法

# 如果是想只获取值，
# for row in ws.iter_rows(min_row=2,max_row=8,min_col=5,max_col=10,values_only=True):
#     print(row)   # 得到的是行列的元组列表。




# 所有行或者所有列
# for row  in ws.rows:
#     print(row)
# for col in ws.columns:
#     print(col)


# 单元格列索引
# print(cell.col_idx)
# print(cell.column)
# 单元格行索引
# print(cell.row)
# 单元格列名
# print(cell.column_letter)
# 单元格的坐标
# print(cell.coordinate)
# 单元格数字类型
# 默认是
# n：数值
# s：字符串
# d：日期时间
# print(cell.data_type)
# 单元格编码格式，默认 utf-8
# print(cell.encoding)
# 是否有样式
# print(cell.has_style)  # 默认样式是 Normal，如果是默认样式，返回False
# 单元格样式
# print(cell.style)
# 单元格样式id
# print(cell.style_id)


# print(ws['A1'].value)
# print(ws.cell(1,2).value)

# 写数据
# for row in range(1,10):
#     ws.append(range(10))
# 10*10的表格


# 使用函数：
# ws["A2"] = "=SUM(1, 10)"   # 这个函数实际上是excel的函数，所有要原样进行复制过来。
# wb.save('shuju.xlsx')


# 单元格合并和拆分
# ws.merge_cells('A2:A9')
# 或者是： ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=4)
# wb.save('shuju.xlsx')


# 插入图片
# from openpyxl.drawing.image import Image
# img=Image('help.jpg')
# ws.add_image(img,'A10')
# wb.save('shuju.xlsx')


'''
性能优化：
openpyxl的内存比是50，内存中的文件大小是实际文件大小的50倍。同时对CPU要求也很高，目的是为了并行处理。

优化的点：
1、设置工作簿的尺寸，将范围限制在有数据的范围内
2、只读模式，如果是要读取cell的数值，可以使用
load_workbook(filename='large_file.xlsx', read_only=True)使用只读模式，这样可以节省大量的内存空间。

3、只写模式。
'''


# 行/列的删除或者插入、移动范围内的单元格
ws.insert_rows(1)
ws.insert_cols(1)
ws.delete_rows(1)
ws.delete_cols(2)
ws.move_range("A1:B2",rows=-3,cols=3) # 这个区域内向上移动3行，向右移动3列。 正是下右，负是上左


'''
创建图表：
可以生成简单的图表：
面积图
条形图
柱状图
气泡图
旭日图
'''



'''
样式：可以应用以下几个方面：
Font：设置字体大小、颜色、下划线等等
PatternFill：设置图案或者颜色渐变
Border：设置单元格的边框
Alignment：单元格对齐
Protection：保护工作表

下面是默认值
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
>>> font = Font(name='Calibri',
...                 size=11,
...                 bold=False,
...                 italic=False,
...                 vertAlign=None,
...                 underline='none',
...                 strike=False,
...                 color='FF000000')
>>> fill = PatternFill(fill_type=None,
...                 start_color='FFFFFFFF',
...                 end_color='FF000000')
>>> border = Border(left=Side(border_style=None,
...                           color='FF000000'),
...                 right=Side(border_style=None,
...                            color='FF000000'),
...                 top=Side(border_style=None,
...                          color='FF000000'),
...                 bottom=Side(border_style=None,
...                             color='FF000000'),
...                 diagonal=Side(border_style=None,
...                               color='FF000000'),
...                 diagonal_direction=0,
...                 outline=Side(border_style=None,
...                              color='FF000000'),
...                 vertical=Side(border_style=None,
...                               color='FF000000'),
...                 horizontal=Side(border_style=None,
...                                color='FF000000')
...                )
>>> alignment=Alignment(horizontal='general',
...                     vertical='bottom',
...                     text_rotation=0,
...                     wrap_text=False,
...                     shrink_to_fit=False,
...                     indent=0)
>>> number_format = 'General'
>>> protection = Protection(locked=True,
...                         hidden=False)
>>>

使用内置样式（Ps：以下注释由译者根据office365中文版进行添加）
该规范（specification）包括一些可以使用的内置样式。不幸的是，这些样式的名称以其本地化形式存储。openpyxl 仅会识别英文名称，并且只能与此处的文字完全一样。
‘Normal’ # 无样式
数字格式:
‘Comma’ # 千位分隔，保留两位小数‘Warning Text’
‘Comma [0]’ # 千位分隔，不保留小数
‘Currency’ # 货币，保留两位小数
‘Currency [0]’ # 货币，不保留小数
‘Percent’ # 百分比
Informative: 
‘Calculation’ # 计算
‘Total’ # 汇总
‘Note’ # 注释
‘Warning Text’ # 警告文本
‘Explanatory Text’ # 解释性文本
文字样式:
‘Title’ # 标题
‘Headline 1’ # 标题1
‘Headline 2’ # 标题2
‘Headline 3’ # 标题3
‘Headline 4’ # 标题4
‘Hyperlink’ # 超链接
‘Followed Hyperlink’ # 已访问的超链接
‘Linked Cell’ # 链接单元格
Comparisons:
‘Input’ # 输入
‘Output’ # 输出
‘Check Cell’ # 检查单元格
‘Good’ # 好
‘Bad’ # 坏
‘Neutral’ # 始终
高亮:
‘Accent1’ # 着色1
‘20 % - Accent1’
‘40 % - Accent1’
‘60 % - Accent1’
‘Accent2’ # 着色2
‘20 % - Accent2’
‘40 % - Accent2’
‘60 % - Accent2’
‘Accent3’ # 着色3
‘20 % - Accent3’
‘40 % - Accent3’
‘60 % - Accent3’
‘Accent4’ # 着色4
‘20 % - Accent4’
‘40 % - Accent4’
‘60 % - Accent4’
‘Accent5’ # 着色5
‘20 % - Accent5’
‘40 % - Accent5’
‘60 % - Accent5’
‘Accent6’ # 着色6
‘20 % - Accent6’
‘40 % - Accent6’
‘60 % - Accent6’
‘Pandas’ # 好像是自定义的


'''


# 筛选和排序：
# ws.auto_filter.ref='A1:W9'  #先指定范围
# ws.auto_filter.add_filter_column(0, ["Kiwi", "Apple", "Mango"])
# ws.auto_filter.add_sort_condition("B2:B15")