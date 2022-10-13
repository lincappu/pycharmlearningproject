# /usr/bin/env python3
# -*- coding: utf-8 -*-
#
# Copyright (C) 2017-2020  The Project X-Ray Authors.
#
# Use of this source code is governed by a ISC-style
# license that can be found in the LICENSE file or at

import openpyxl


'''
层级：
workbook  工作簿，就是一个xlsl文件
worksheet  工作表  文件里面的一个表
rows  一行
column  一列
cell  一个单元格
'''


# 一、文件操作
# 新建操作
# wb=openpyxl.Workbook('a.xlsx')
# wb.create_sheet('test')
# print(wb.sheetnames)
# ws=wb.active  # 当前活动的表，就是第一个表。
# wb.save('a.xlsx')
#  打开操作
wb = openpyxl.load_workbook('shuju.xlsx')
print(wb.sheetnames)
print(wb.encoding)
ws=wb['FL']
print(ws.title)
print(ws)

#  工作簿。工作表、单元格的属性
print(wb.sheetnames)
print(wb.worksheets) # 所有的表名

print(ws.title) # 当前表名
print(wb.read_only) # 判断是否只读
print(wb.encoding)
print(wb.properties)  # 工作簿元数据

print(ws.max_row)
print(ws.max_column)


# 工作簿、表的操作。
# 复制工作表
# target=wb.copy_worksheet(ws)
# print(target)
# target.title='FLS'
# wb.save('shuju.xlsx')



# 读取单元格数据
# cells=ws['A:B]
# cells=ws['5:6:7:8:9:10:11:12:']

# # 获取区域的情况
# for row in ws.iter_rows(min_row=2,max_row=10,min_col=2,max_col=3):
#     print(row)
#     for cell in row:
#         print(cell.value)

# ws.iter_cols()


# 所有行：
# for row  in ws.rows:
#     print(row)

# 单元格列索引
# print(cell.col_idx)
# print(cell.column)
# 单元格行索引
# print(cell.row)
# 单元格列名
# print(cell.column_letter)
# 单元格的坐标
# print(cell.coordinate)
# 单元格数字类型
# 默认是
# n：数值
# s：字符串
# d：日期时间
# print(cell.data_type)
# 单元格编码格式，默认 utf-8
# print(cell.encoding)
# 是否有样式
# print(cell.has_style)  # 默认样式是 Normal，如果是默认样式，返回False
# 单元格样式
# print(cell.style)
# 单元格样式id
# print(cell.style_id)


# print(ws['A1'].value)
# print(ws.cell(1,2).value)

# 写数据
for row in range(1,10):
    ws.append(range(10))
# 10*10的表格


# 函数sum ，

# 筛选机制：
ws.auto_filter.ref='A1"W9'  #先指定范围